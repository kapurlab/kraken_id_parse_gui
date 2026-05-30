import datetime as _dt
import json
import os
import platform
import re
import shutil
import subprocess
from pathlib import Path
from typing import Any, Dict, Iterable, List, Optional


SCHEMA_VERSION = "0.1"

_ARTIFACT_EXTENSIONS = {
    ".csv": "table",
    ".fa": "sequence",
    ".fasta": "sequence",
    ".gz": "compressed",
    ".html": "html",
    ".htm": "html",
    ".json": "json",
    ".log": "log",
    ".pdf": "pdf",
    ".png": "figure",
    ".txt": "text",
    ".vcf": "variant",
    ".xls": "spreadsheet",
    ".xlsx": "spreadsheet",
}


def _now_iso() -> str:
    return _dt.datetime.now(_dt.timezone.utc).isoformat()


def _today_label() -> str:
    today = _dt.datetime.now().astimezone()
    return f"{today.strftime('%B')} {today.day}, {today.year}"


def _version(command: Iterable[str]) -> Optional[str]:
    argv = list(command)
    if not argv or not shutil.which(argv[0]):
        return None
    try:
        result = subprocess.run(
            argv,
            check=False,
            capture_output=True,
            text=True,
            timeout=8,
        )
    except (OSError, subprocess.TimeoutExpired):
        return None
    text = (result.stdout or result.stderr or "").strip().splitlines()
    return text[0].strip() if text else None


def _git_commit(repo: Path) -> Optional[str]:
    try:
        result = subprocess.run(
            ["git", "-C", str(repo), "rev-parse", "--short", "HEAD"],
            check=False,
            capture_output=True,
            text=True,
            timeout=5,
        )
    except (OSError, subprocess.TimeoutExpired):
        return None
    commit = result.stdout.strip()
    return commit or None


def _artifact_kind(path: Path) -> str:
    if path.name.endswith("_krona.html"):
        return "krona"
    if path.name.endswith("_report.pdf"):
        return "report_pdf"
    if path.name.endswith("_stats.xlsx"):
        return "summary_spreadsheet"
    return _ARTIFACT_EXTENSIONS.get(path.suffix.lower(), "artifact")


def collect_artifacts(output_dir: Path) -> List[Dict[str, Any]]:
    artifacts: List[Dict[str, Any]] = []
    for path in sorted(output_dir.rglob("*")):
        if not path.is_file():
            continue
        if ".report_assets" in path.parts:
            continue
        if path.name in {"run_manifest.json", "report.html"}:
            continue
        try:
            rel = path.relative_to(output_dir)
            stat = path.stat()
        except OSError:
            continue
        artifacts.append(
            {
                "name": path.name,
                "path": str(rel),
                "kind": _artifact_kind(path),
                "size_bytes": stat.st_size,
            }
        )
    return artifacts


def _first_stats_row(output_dir: Path) -> Dict[str, Any]:
    try:
        from openpyxl import load_workbook
    except ImportError:
        return {}

    stats_files = sorted(output_dir.glob("*_stats.xlsx"))
    if not stats_files:
        return {}
    try:
        workbook = load_workbook(stats_files[-1], data_only=True, read_only=True)
        sheet = workbook.active
        rows = list(sheet.iter_rows(values_only=True))
    except Exception:
        return {}
    if len(rows) < 2:
        return {}
    headers = [str(h) if h is not None else "" for h in rows[0]]
    values = rows[1]
    return {headers[i]: values[i] for i in range(min(len(headers), len(values))) if headers[i]}


def _parse_classification_summary(summary: Any) -> List[Dict[str, Any]]:
    if not summary:
        return []
    rows = []
    for part in str(summary).split(";"):
        part = part.strip()
        match = re.match(r"(.+?)\s+\(([^)]+)\):\s+(.+)", part)
        if match:
            rows.append({"name": match.group(1), "level": match.group(2), "percent": match.group(3)})
    return rows


def _kraken_line_to_row(line: str) -> Optional[Dict[str, Any]]:
    parts = line.rstrip("\n").split("\t")
    if len(parts) < 6:
        return None
    return {
        "percent": parts[0].strip(),
        "reads": parts[1].strip(),
        "direct_reads": parts[2].strip(),
        "level": parts[3].strip(),
        "taxon_id": parts[4].strip(),
        "name": parts[5].strip(),
    }


def _target_taxonomy_rows(output_dir: Path, names: List[str]) -> List[Dict[str, Any]]:
    report_files = sorted((output_dir / "kraken").glob("*_reportkraken.txt"))
    if not report_files:
        return []
    wanted = {name.strip() for name in names if name.strip()}
    rows = []
    try:
        for line in report_files[-1].read_text(encoding="utf-8", errors="replace").splitlines():
            row = _kraken_line_to_row(line)
            if not row:
                continue
            if row["name"] in wanted:
                rows.append(row)
    except OSError:
        return []
    return rows


def _read_bracken_rows(output_dir: Path, limit: int = 8) -> List[Dict[str, Any]]:
    try:
        from openpyxl import load_workbook
    except ImportError:
        return []

    files = sorted((output_dir / "kraken").glob("*-bracken.xlsx"))
    if not files:
        return []
    try:
        workbook = load_workbook(files[-1], data_only=True, read_only=True)
        rows = list(workbook.active.iter_rows(values_only=True))
    except Exception:
        return []
    if len(rows) < 2:
        return []
    headers = [str(h) if h is not None else "" for h in rows[0]]
    out = []
    for values in rows[1 : limit + 1]:
        row = {headers[i]: values[i] for i in range(min(len(headers), len(values))) if headers[i]}
        if row:
            out.append(row)
    return out


def _read_blast_rows(output_dir: Path, limit: int = 12) -> List[Dict[str, Any]]:
    files = sorted(output_dir.glob("*_blast_summary.txt"))
    rows = []
    for path in files:
        try:
            lines = path.read_text(encoding="utf-8", errors="replace").splitlines()
        except OSError:
            continue
        for line in lines[:limit]:
            parts = line.split("\t", 2)
            if len(parts) == 3:
                rows.append(
                    {
                        "source": path.name,
                        "nt_base_count": parts[0],
                        "contigs": parts[1],
                        "description": parts[2],
                    }
                )
        if rows:
            break
    return rows


def _fasta_summary(output_dir: Path) -> List[Dict[str, Any]]:
    try:
        from Bio import SeqIO
    except ImportError:
        return []
    rows = []
    for path in sorted(output_dir.glob("*.fasta")):
        lengths = []
        try:
            for record in SeqIO.parse(path, "fasta"):
                lengths.append(len(record.seq))
        except Exception:
            continue
        if lengths:
            rows.append(
                {
                    "name": path.name,
                    "count": len(lengths),
                    "total_length": sum(lengths),
                    "longest": max(lengths),
                }
            )
    return rows


def _linked_artifacts(output_dir: Path, kind: str) -> List[Dict[str, Any]]:
    return [artifact for artifact in collect_artifacts(output_dir) if artifact["kind"] == kind]


def _coverage_figures(output_dir: Path) -> List[Dict[str, Any]]:
    figures = []
    asset_dir = output_dir / ".report_assets"
    pdftoppm = shutil.which("pdftoppm")

    coverage_pdfs = sorted(output_dir.glob("*-coverage_graph.pdf"), key=lambda path: path.stat().st_mtime)
    # Legacy LaTeX reports surfaced the final selected coverage graph in the
    # report body. Earlier exploratory coverage graphs can remain in the run
    # directory across reruns, so keep the newest graph in the printable report.
    if len(coverage_pdfs) > 1:
        coverage_pdfs = coverage_pdfs[-1:]

    for pdf in coverage_pdfs:
        preview = pdf.with_suffix(".png")
        if not preview.exists() and pdftoppm:
            try:
                asset_dir.mkdir(exist_ok=True)
                preview = asset_dir / f"{pdf.stem}.png"
                subprocess.run(
                    [pdftoppm, "-png", "-f", "1", "-singlefile", "-r", "150", str(pdf), str(preview.with_suffix(""))],
                    check=False,
                    capture_output=True,
                    text=True,
                    timeout=30,
                )
            except (OSError, subprocess.TimeoutExpired):
                preview = pdf.with_suffix(".png")

        figures.append(
            {
                "name": pdf.name,
                "path": str(pdf.relative_to(output_dir)),
                "preview_path": str(preview.relative_to(output_dir)) if preview.exists() else None,
            }
        )
    return figures


def _coverage_stats(output_dir: Path) -> List[Dict[str, Any]]:
    stats_files = sorted(output_dir.glob("*-coverage_stats.json"), key=lambda path: path.stat().st_mtime)
    if not stats_files:
        return []
    try:
        data = json.loads(stats_files[-1].read_text(encoding="utf-8"))
    except (OSError, json.JSONDecodeError):
        return []
    return data if isinstance(data, list) else []


def collect_legacy_report(output_dir: Path) -> Dict[str, Any]:
    stats = _first_stats_row(output_dir)
    classifications = _parse_classification_summary(stats.get("Classification Summary"))
    taxonomy = _target_taxonomy_rows(output_dir, [row["name"] for row in classifications])
    return {
        "fastq": [
            {
                "read": "R1",
                "filename": stats.get("FASTQ_R1"),
                "file_size": stats.get("R1 File Size"),
                "read_count": stats.get("R1 Read Count"),
                "q30": stats.get("R1 Passing Q30"),
                "mean_score": stats.get("R1 Read Quality Ave"),
                "avg_length": stats.get("R1 Ave Length"),
            },
            {
                "read": "R2",
                "filename": stats.get("FASTQ_R2"),
                "file_size": stats.get("R2 File Size"),
                "read_count": stats.get("R2 Read Count"),
                "q30": stats.get("R2 Passing Q30"),
                "mean_score": stats.get("R2 Read Quality Ave"),
                "avg_length": stats.get("R2 Ave Length"),
            },
        ],
        "target": {
            "taxon": stats.get("Target Taxon"),
            "taxon_id": stats.get("Taxon ID"),
            "total_input_reads": stats.get("Total Input Reads"),
            "extracted_reads": stats.get("Extracted Reads"),
            "extraction_rate": stats.get("Extraction Rate (%)"),
        },
        "classification_summary": classifications,
        "target_taxonomy": taxonomy,
        "bracken_species": _read_bracken_rows(output_dir),
        "assembly": {
            "contig_count": stats.get("Contig count"),
            "contig_bins": stats.get("Contig length counts <|301-999bp|>"),
            "longest_contig": stats.get("Longest contig"),
            "total_length": stats.get("Total length"),
            "n50": stats.get("N50"),
            "mean_coverage": stats.get("FASTQ calculated mean coverage"),
        },
        "top_blast_columns": [
            {"label": key, "value": value}
            for key, value in stats.items()
            if key.startswith("Top BLAST") and value
        ],
        "blast_rows": _read_blast_rows(output_dir),
        "fasta_summary": _fasta_summary(output_dir),
        "coverage_figures": _coverage_figures(output_dir),
        "coverage_stats": _coverage_stats(output_dir),
        "krona": _linked_artifacts(output_dir, "krona"),
    }


def build_run_manifest(
    *,
    sample_id: str,
    status: str,
    inputs: Dict[str, Optional[str]],
    parameters: Dict[str, Optional[str]],
    output_dir: Path,
    started_at: Optional[_dt.datetime] = None,
    warnings: Optional[List[str]] = None,
    sections: Optional[List[Dict[str, Any]]] = None,
    repo_root: Optional[Path] = None,
) -> Dict[str, Any]:
    repo_root = repo_root or Path(__file__).resolve().parents[2]
    started = started_at.isoformat() if started_at else None
    finished = _now_iso()
    databases = []
    if parameters.get("kraken_db"):
        databases.append({"name": Path(str(parameters["kraken_db"])).name, "path": parameters["kraken_db"], "type": "kraken"})
    if parameters.get("blast_db"):
        databases.append({"name": Path(str(parameters["blast_db"])).name, "path": parameters["blast_db"], "type": "blast"})

    return {
        "schema_version": SCHEMA_VERSION,
        "report_date": _today_label(),
        "sample_id": sample_id,
        "run_id": output_dir.name,
        "status": status,
        "warnings": warnings or [],
        "inputs": inputs,
        "parameters": parameters,
        "software": {
            "pipeline_git_commit": _git_commit(repo_root),
            "python": platform.python_version(),
            "kraken2": _version(["kraken2", "--version"]),
            "bracken": _version(["bracken", "-v"]),
            "blastn": _version(["blastn", "-version"]),
            "tectonic": _version(["tectonic", "--version"]),
        },
        "databases": databases,
        "environment": {
            "hostname": platform.node(),
            "user": os.environ.get("USER"),
            "cwd": str(output_dir),
        },
        "timing": {
            "started_at": started,
            "finished_at": finished,
        },
        "legacy_report": collect_legacy_report(output_dir),
        "sections": sections or [],
        "artifacts": collect_artifacts(output_dir),
    }


def write_manifest(manifest: Dict[str, Any], output_dir: Path) -> Path:
    path = output_dir / "run_manifest.json"
    path.write_text(json.dumps(manifest, indent=2, sort_keys=True) + "\n", encoding="utf-8")
    return path
